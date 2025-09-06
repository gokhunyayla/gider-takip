from flask import render_template, redirect, url_for, flash, request, jsonify, Response
from flask_login import login_required, current_user
from datetime import datetime
from app import db
from app.models import ExpenseRecord, ExpenseCategory
from . import main_bp
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

@main_bp.route('/expenses/add', methods=['POST'])
@login_required
def add_expense():
    """Add new expense record"""
    try:
        # Get form data
        category_id = request.form.get('category_id')
        year = request.form.get('year', type=int)
        month = request.form.get('month', type=int)
        day = request.form.get('day', type=int)
        amount = request.form.get('amount', type=float)
        description = request.form.get('description')
        invoice_number = request.form.get('invoice_number')
        supplier_name = request.form.get('supplier_name')
        payment_status = 'paid'  # Tüm kayıtlar otomatik ödendi
        
        # Create due_date from year, month, day if day is provided
        due_date = None
        if day and year and month:
            try:
                due_date = datetime(year, month, day).date()
            except ValueError:
                # Invalid date combination
                pass
        
        # Validate required fields
        if not all([category_id, year, month, amount]):
            flash('Lütfen tüm zorunlu alanları doldurun.', 'danger')
            return redirect(url_for('main.expenses'))
        
        # Create expense record
        expense = ExpenseRecord(
            company_id=current_user.company_id,
            category_id=category_id,
            year=year,
            month=month,
            amount=amount,
            description=description,
            invoice_number=invoice_number,
            supplier_name=supplier_name,
            payment_status=payment_status,
            due_date=due_date,  # Gün bilgisi burada saklanıyor
            user_id=current_user.id
        )
        
        # Set payment date (always now since all expenses are paid)
        expense.payment_date = datetime.now().date()
        
        db.session.add(expense)
        db.session.commit()
        
        flash('Gider kaydı başarıyla eklendi.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Hata oluştu: {str(e)}', 'danger')
    
    # Check if request came from dashboard
    if request.referrer and 'dashboard' in request.referrer:
        return redirect(url_for('main.dashboard'))
    
    return redirect(url_for('main.expenses'))

@main_bp.route('/expenses/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_expense(id):
    """Edit expense record"""
    expense = ExpenseRecord.query.filter_by(
        id=id,
        company_id=current_user.company_id
    ).first_or_404()
    
    if request.method == 'POST':
        # Update expense
        expense.category_id = request.form.get('category_id', expense.category_id)
        expense.year = request.form.get('year', expense.year, type=int)
        expense.month = request.form.get('month', expense.month, type=int)
        expense.amount = request.form.get('amount', expense.amount, type=float)
        expense.description = request.form.get('description', expense.description)
        expense.invoice_number = request.form.get('invoice_number', expense.invoice_number)
        expense.supplier_name = request.form.get('supplier_name', expense.supplier_name)
        expense.payment_status = 'paid'  # Always paid
        
        # Update due_date with day if provided
        day = request.form.get('day', type=int)
        if day and expense.year and expense.month:
            try:
                expense.due_date = datetime(expense.year, expense.month, day).date()
            except ValueError:
                # Invalid date combination
                pass
        
        # Ensure payment date is set
        if not expense.payment_date:
            expense.payment_date = datetime.now().date()
        
        db.session.commit()
        flash('Gider kaydı güncellendi.', 'success')
        return redirect(url_for('main.expenses'))
    
    # GET request - return JSON for modal
    return jsonify({
        'id': expense.id,
        'category_id': expense.category_id,
        'year': expense.year,
        'month': expense.month,
        'day': expense.due_date.day if expense.due_date else None,
        'amount': float(expense.amount),
        'description': expense.description,
        'invoice_number': expense.invoice_number,
        'supplier_name': expense.supplier_name,
        'payment_status': expense.payment_status,
        'due_date': expense.due_date.strftime('%Y-%m-%d') if expense.due_date else None
    })

@main_bp.route('/expenses/<int:id>/delete', methods=['POST'])
@login_required
def delete_expense(id):
    """Delete expense record"""
    expense = ExpenseRecord.query.filter_by(
        id=id,
        company_id=current_user.company_id
    ).first_or_404()
    
    db.session.delete(expense)
    db.session.commit()
    
    flash('Gider kaydı silindi.', 'info')
    return redirect(url_for('main.expenses'))

@main_bp.route('/expenses/export')
@login_required
def export_expenses():
    """Export expenses to Excel file"""
    # Get filter parameters from request
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    category_id = request.args.get('category_id', type=int)
    
    # Build query
    query = ExpenseRecord.query.filter_by(company_id=current_user.company_id)
    
    if year:
        query = query.filter_by(year=year)
    if month:
        query = query.filter_by(month=month)
    if category_id:
        query = query.filter_by(category_id=category_id)
    
    # Get expenses sorted by newest first
    expenses = query.order_by(
        ExpenseRecord.year.desc(),
        ExpenseRecord.month.desc(),
        ExpenseRecord.created_at.desc()
    ).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Giderler"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Month names
    month_names = {
        1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan",
        5: "Mayıs", 6: "Haziran", 7: "Temmuz", 8: "Ağustos",
        9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"
    }
    
    # Add headers
    headers = ["Tarih", "Kategori", "Açıklama", "Tedarikçi", "Fatura No", "Tutar (₺)", "Durum"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    for row, expense in enumerate(expenses, 2):
        # Format date
        if expense.due_date:
            date_str = f"{expense.due_date.day:02d} {month_names[expense.month]} {expense.year}"
        else:
            date_str = f"{month_names[expense.month]} {expense.year}"
        
        ws.cell(row=row, column=1, value=date_str).border = border
        ws.cell(row=row, column=2, value=expense.category.name if expense.category else "").border = border
        ws.cell(row=row, column=3, value=expense.description or "").border = border
        ws.cell(row=row, column=4, value=expense.supplier_name or "").border = border
        ws.cell(row=row, column=5, value=expense.invoice_number or "").border = border
        
        # Amount with formatting
        amount_cell = ws.cell(row=row, column=6, value=float(expense.amount))
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = border
        
        # Status
        status = "Ödendi" if expense.payment_status == "paid" else "Bekliyor"
        ws.cell(row=row, column=7, value=status).border = border
    
    # Add summary row
    summary_row = len(expenses) + 3
    ws.cell(row=summary_row, column=5, value="TOPLAM:").font = Font(bold=True)
    total_cell = ws.cell(row=summary_row, column=6, value=sum(float(e.amount) for e in expenses))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'
    
    # Adjust column widths
    column_widths = [20, 25, 40, 25, 15, 15, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename_parts = ["giderler"]
    if year:
        filename_parts.append(str(year))
    if month:
        filename_parts.append(month_names[month])
    filename = f"{'_'.join(filename_parts)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={filename}',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )

